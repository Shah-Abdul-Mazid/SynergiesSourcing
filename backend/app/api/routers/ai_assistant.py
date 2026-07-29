from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Body
from motor.motor_asyncio import AsyncIOMotorDatabase
from typing import List, Optional
from app.core.database import get_db
import app.crud.erp_crud as crud
from app.schemas.erp_schemas import (
    AIAssistantRequest, AIAssistantResponse, RAGDocumentOut,
    EmailLogCreate, EmailLogOut, AITaskCreate, AITaskOut
)
from app.services.ai_service import AIService
from datetime import datetime
import io

try:
    import PyPDF2
except Exception:
    PyPDF2 = None

try:
    import docx
except Exception:
    docx = None

try:
    import openpyxl
except Exception:
    openpyxl = None

router = APIRouter(prefix="/assistant", tags=["ai_assistant"])
ai_service = AIService()

def _extract_text_from_file(filename: str, file_bytes: bytes) -> str:
    ext = filename.split(".")[-1].lower()
    text_content = ""

    try:
        if ext == "pdf" and PyPDF2 is not None:
            reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
            for page in reader.pages:
                text_content += (page.extract_text() or "") + "\n"
        
        elif ext == "docx" and docx is not None:
            doc = docx.Document(io.BytesIO(file_bytes))
            for para in doc.paragraphs:
                text_content += para.text + "\n"
        
        elif ext in ["xlsx", "xls"] and openpyxl is not None:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                text_content += f"--- Sheet: {sheet} ---\n"
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_str = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_str:
                        text_content += row_str + "\n"
        
        elif ext == "csv":
            text_content = file_bytes.decode("utf-8", errors="ignore")
        
        else:
            text_content = file_bytes.decode("utf-8", errors="ignore")
            
        return text_content.strip()
    except Exception as e:
        raise ValueError(f"Failed to extract text from {ext} file: {str(e)}")


@router.post("/chat", response_model=AIAssistantResponse)
async def chat_with_erp_assistant(payload: AIAssistantRequest, db: AsyncIOMotorDatabase = Depends(get_db)):
    """Chat assistant with context injections from database based on selection."""
    try:
        context = {}
        if payload.context_type in ["orders", "all"]:
            orders = await crud.get_all_purchase_orders(db)
            context["active_orders"] = [
                {"id": o.order_id, "buyer": o.buyer_name, "item": o.item, "qty": o.quantity, "status": o.status, "risk": o.risk_level}
                for o in orders
            ]
        
        if payload.context_type in ["inventory", "all"]:
            inventory = await crud.get_all_inventory(db)
            context["materials_inventory"] = [
                {"item": i.item, "quantity": i.quantity, "unit": i.unit}
                for i in inventory
            ]
        
        if payload.context_type in ["suppliers", "all"]:
            suppliers = await crud.get_all_suppliers(db)
            context["registered_suppliers"] = [
                {"id": s.supplier_id, "name": s.name, "email": s.contact_email}
                for s in suppliers
            ]
            
        if payload.context_type in ["qc", "all"]:
            qc_logs = await crud.get_all_qc_logs(db)
            context["quality_inspection_history"] = [
                {"log_id": q.log_id, "order": q.order_id, "defect": q.defect_type, "status": q.status}
                for q in qc_logs
            ]

        response_text = await ai_service.chat_with_erp(payload.message, context)
        
        actions = []
        lower_resp = response_text.lower()
        if "reorder" in lower_resp or "safety stock" in lower_resp:
            actions.append("Order Restock replenishments")
        if "warning" in lower_resp or "high risk" in lower_resp:
            actions.append("Run Disruption Simulator")
        if "defect" in lower_resp or "qc" in lower_resp:
            actions.append("Run Fabric Scan")
            
        return {
            "response": response_text,
            "actions": actions
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rag/upload", response_model=RAGDocumentOut)
async def upload_rag_document(
    file: UploadFile = File(...),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Upload PDF/DOCX/Excel/CSV contract, invoice, or tech pack and index in local database RAG library."""
    try:
        file_bytes = await file.read()
        extracted_text = _extract_text_from_file(file.filename, file_bytes)
        
        if not extracted_text:
            raise HTTPException(status_code=400, detail="Document contains no readable text content.")
            
        doc_data = {
            "filename": file.filename,
            "doc_type": file.filename.split(".")[-1].upper(),
            "content_text": extracted_text,
            "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        doc = await crud.create_rag_document(db, doc_data)
        return doc
    except ValueError as val_err:
        raise HTTPException(status_code=400, detail=str(val_err))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database write failure: {str(e)}")


@router.get("/rag/documents", response_model=List[RAGDocumentOut])
async def list_rag_documents(db: AsyncIOMotorDatabase = Depends(get_db)):
    """List all indexed RAG documents."""
    try:
        return await crud.get_all_rag_documents(db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rag/query")
async def query_rag_knowledge_base(
    question: str = Body(..., embed=True),
    db: AsyncIOMotorDatabase = Depends(get_db)
):
    """Answer user question using ONLY information from uploaded documents (RAG)."""
    try:
        docs = await crud.get_all_rag_documents(db)
        if not docs:
            return {
                "response": "No documents uploaded. Please upload a contract, invoice, or tech pack to the RAG knowledge base first.",
                "sources": []
            }
        
        combined_text = ""
        sources = []
        for d in docs:
            combined_text += f"\n=== START FILE: {d.filename} ===\n{d.content_text}\n=== END FILE ===\n"
            sources.append(d.filename)
            
        if len(combined_text) > 12000:
            combined_text = combined_text[:12000] + "... (content truncated for prompt size)"
            
        answer = await ai_service.answer_from_documents(question, combined_text)
        return {
            "response": answer,
            "sources": sources
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
